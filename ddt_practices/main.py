import unittest
from openpyxl import load_workbook
from ddt import ddt, data, unpack

def get_excel_data(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]  # 获取表头
    data_list = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳过表头
        data_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_list.append(data_dict)

    return data_list

#关键字配合管理解释器
def filter_data_by_case(data_list, case_flag):
    return [data for data in data_list if data.get('TestCaseFlag') == case_flag]


@ddt
class TestOneLearAndUseDdt(unittest.TestCase):
    test_data = get_excel_data('./testdata_ddt.xlsx')
    #case1数据执行测试1
    @data(*filter_data_by_case(test_data,'case1'))
    @unpack
    def testOne(self, **case):
        print(case)

    # case2数据执行测试2
    @data(*filter_data_by_case(test_data, 'case2'))
    @unpack
    def testTwo(self, **case):
        print(case['Method'])
        #print(case)
if __name__ == '__main__':
    unittest.main()
