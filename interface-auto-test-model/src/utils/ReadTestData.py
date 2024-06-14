# -*- coding: utf-8 -*-
# @Project : 接口自动化测试
# @File    : ReadTestData.py
# @Author  : 朱宽
# @Time    : 2021/3/31 15:36
# @Software: Win10 / Python3 / Pycharm

# '''
#     命令行执行某python文件时
#     用于将项目的路径添加到临时空间中（即工作目录），便于寻找该项目下的其他模块
# '''
# import sys
# import os
# item_path=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# sys.path.append(item_path)

from openpyxl import load_workbook  # 读取excel中的数据
from config import globalConfig  # 获取需要读取“测试数据”的路径

from copy import deepcopy
import re

class ReadExcel():
    '''
    1、read_data()，遍历每一行数据并放到List,并返回list，使用列表套列表的形式，每一行为一个内层列表，无数行存储在一个大列表中
    2、output_data_dict()，将数据中的内层列表转换为字典存储，外层依然是列表
    3、test_data()，将AssertInfo,SendData等字段对应的值转换为 dict 形式
    '''
    def __init__(self,file_name):
        '''
        初始化，为读取excel做准备
        '''
        self.filePath= globalConfig.test_data_path + file_name

        self.wb = load_workbook(self.filePath)
        # 获取所有sheet页名字
        self.SheetNames = self.wb.sheetnames
        self.sheet = self.wb[self.SheetNames[0]]  # self.wb[SheetName]
        self.MaxRowNum = self.sheet.max_row
        self.MaxColNum=self.sheet.max_column


    def read_data(self):
        '''
        2、遍历每一行数据并放到List,并返回list，使用列表套列表的形式，每一行为一个内层列表，无数行存储在一个大列表中
        :return:
        '''
        dataList = []
        for line in list(self.sheet.rows)[0:]:  # 获取excel中每一行的数据。rows这个函数 用来按行读取；columns这个函数 用来按列读取。
            tmpList = []
            for i in range(0, self.MaxColNum):  # 指定获取每一行中的1-12列的数据，不需要 12列后面的数据。
                tmpList.append(line[i].value)
            dataList.append(tmpList)
        return dataList

    def output_data_dict(self):
        '''
        3、将数据中的内层列表转换为字典存储，外层依然是列表。并返回
        形如：
        [
            {'me':'zk','you':'Bill'},
            {'me1':'zk1','you2':'Bill2'}
        ]
        :return:外列表，内字典。内字典中的 ’send_data‘对应的value值是字符串，需要使用 eval()函数 将其转换成字典才能使用。
        '''
        #print('最大行数：', self.MaxRowNum,'\n最大列数：',self.MaxColNum)
        readDataList = []
        read_data_list = self.read_data()
        # pprint(read_data_list)
        for i in range(1, len(read_data_list)):  # 外层列表循环
            temDict = {}
            for j in range(0, len(read_data_list[0])):  # 内层列表循环
                '''
                下面的字典赋值语句，将excel中的第一行中的字段名，与每一列中的数据一一对应起来，形成key-value这样的字典。
                '''
                temDict[read_data_list[0][j]] = read_data_list[i][j]
            readDataList.append(temDict)  # 将每一行数据 与 字段名 形成的 字典 作为 列表的值 存储起来。
        return readDataList

    def test_data(self):
        '''
        4、将AssertInfo,SendData等字段对应的值转换为 dict 形式

        *** 关于正则表达式的解释 ***

            re.search("\{.*\}",list['SendData'],re.S)!=None
            解释：
            （1）\{.*\} ，表示在母串中匹配 “{} 中有任意字符” 的子串；\{，表示转义，因为{}在正则表达式中有特殊含义；.* 表示任意字符（通用匹配）
            （2）list['SendData'] 母串
            （3）re.S ，表示使“.”匹配包括换行在内的所有字符；“.” 表示匹配任意字符（除换行符\n）
            （4）当search()有结果时，会返回相应的对象结果；如果没有结果的话，那么就会返回 None

        :return: test_data，作为读取的最终输出；数据结构 [{{}},{{}}]
        '''
        readDataList=self.output_data_dict()
        test_data=deepcopy(readDataList)

        for i in range(0,len(readDataList)):
            list=readDataList[i]

            '''
                下面这些if判断都是实现的同样功能
                （1）当某值不为 None 
                （2）当该值中存在 “{任意字符}” 这样的字符；详情请看方法注释
                （3）同时符合（1）和（2）中条件时则执行if下面的代码
                （4）将该值使用 eval() 函数进行转换后，并对相应的key重新赋值
            '''
            for key in list.keys():
                if list[key]!=None and re.search("\{.*\}",str(list[key]),re.S)!=None:
                    field_data=eval(list[key])
                    test_data[i][key]=field_data
                elif list[key]!=None and re.search("\[.*\]",str(list[key]),re.S)!=None:
                    field_data = eval(list[key])
                    test_data[i][key] = field_data

        return test_data



if __name__=="__main__":
    import json
    test=ReadExcel('testdata.xlsx')
    data_all=test.test_data()

    print(json.dumps(data_all,indent=4,ensure_ascii=False)) # 按照json格式输出
    print(type(data_all))
    print(type(data_all[0]))
    print(type(data_all[0]['SendData']))



