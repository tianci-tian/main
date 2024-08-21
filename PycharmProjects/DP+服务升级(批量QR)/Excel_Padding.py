
from openpyxl import load_workbook


#适配excel表格第一行行数,第一列列数,全部按照第一行内容填充

def Excel_Padding(file):
    #1.打开excel文件
    workbook = load_workbook(file)
    #2.通过sheet名称获取表格   只有一张表格的时候，可以直接 active
    # sheet1 = workbook['Sheet1']
    sheet1 = workbook.active
    # 3.获取表格内某个格子的数据
    #cell1 = sheet1['A1']
    # 获取最大行数
    row_count = sheet1.max_row
    # 获取最大列数
    max_column = sheet1.max_column
    #填充
    first_row_values = [cell.value for cell in sheet1[2]]
    for row in range(2, row_count + 1):
        for col in range(2, max_column + 1):
            sheet1.cell(row=row, column=col).value = first_row_values[col - 1]
    workbook.save(file)

        #test
if __name__ =='__main__':
    excel_path = 'D:\\桌面\\testexcel.xlsx'
    Excel_Padding(excel_path)