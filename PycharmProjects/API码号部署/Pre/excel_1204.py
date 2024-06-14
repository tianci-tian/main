from API码号部署.Pre.encrypt_jar import run_java_jar
from openpyxl import Workbook, load_workbook

'''
读：明文.txt   依据1组数据写100组iccid，使用num_iterations变量控制循环次数
写入：profile_apiTest_45407_002.inp
调用jar包加密，加密数据写入：profile_apiTest_45407_002.txt
写入：profile.xlsx；profile.xlsx中Sheet1 A列展示：iccid  B列展示：iccid码号对应的密文
num_iterations变量控制循环次数；iccid遍历+1
'''


def modify_file_content(file_path, num_iterations):
    for _ in range(num_iterations):
        with open(file_path, 'r') as file:
            lines = file.readlines()

        if len(lines) >= 11:
            line_to_modify = lines[11].strip()
            elements = line_to_modify.split()  # 以空格分割
            if elements:
                plain_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\profile_apiTest_45407_002.inp"
                for i in range(0,100):
                    first_element = int(elements[0])
                    first_element += 1
                    elements[0] = str(first_element)
                    print(elements)
                    lines[i+11] = ' '.join(elements) + '\n' # 更新第11行的内容
                    with open(plain_address, 'w') as profile_file:
                        profile_file.writelines(lines)



            ##调用加密jar包
            plain_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\profile_apiTest_45407_002.inp"
            public_key_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\links_apiTest_publickey.asc"
            output_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\profile_apiTest_45407_002.txt"
            jar_path = "D:\\downloads\\encrypt-1.0-SNAPSHOT.jar"



            data = run_java_jar(plain_address, public_key_address, output_address, jar_path)
            print(data)

            # 删除第11行到第20行
            del lines[11:110]
            # 在删除的位置插入99行空行
            lines[11:11] = ['\n'] * 99
            # 将修改后的内容写回文件
            with open(file_path, 'w') as file:
                file.writelines(lines)
            # 将第11行和第21行的内容互换
            lines[11], lines[110] = lines[110], lines[11]
            # 将修改后的内容写回文件
            with open(file_path, 'w') as file:
                file.writelines(lines)
        '''
            data写入到D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx文件的B2单元格内，第二次循环写入到B3,第三次循环写入到B4，依次类推

        '''
        # 将数据写入 Excel 文件的不同行
        wb = load_workbook('D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx')
        ws = wb.active
        ws.cell(row=_ + 2, column=1, value=str(first_element))
        ws.cell(row=_ + 2, column=2, value=data)
        wb.save('D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx')
        # 将修改后的内容写回文件
        with open(file_path, 'w') as file:
            file.writelines(lines)


file_path = 'D:\\项目\\DP+\\实施部署API\\Pre环境\\明文.txt'
num_iterations = 300
modify_file_content(file_path, num_iterations)
