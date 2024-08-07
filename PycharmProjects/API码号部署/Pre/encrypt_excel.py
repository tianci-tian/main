from PycharmProjects.API码号部署.Pre.encrypt_jar import run_java_jar
from openpyxl import Workbook, load_workbook

'''
读：明文.txt       一组数据1个iccid
写入：profile_apiTest_45407_002.inp
调用jar包加密，加密数据写入：profile_apiTest_45407_002.txt
写入：profile.xlsx；profile.xlsx中Sheet1 A列展示：iccid  B列展示：iccid码号对应的密文
num_iterations变量控制循环次数；iccid遍历+1
'''
def modify_file_content(file_path, num_iterations):
    for _ in range(num_iterations):
        with open(file_path, 'r') as file:
            lines = file.readlines()

        if len(lines) >= 11:
            line_to_modify = lines[11].strip()
            elements = line_to_modify.split()
            if elements:
                first_element = int(elements[0])
                first_element += 1
                elements[0] = str(first_element)
                lines[11] = ' '.join(elements) + '\n'  # 更新第11行的内容
            ##调用pgp加密jar包
            plain_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\profile_apiTest_45407_002.inp"
            public_key_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\links_apiTest_publickey.asc"
            Private_key_address = "D:\桌面\TTCPrivateKeys.asc"
            passwords = '123456'
            output_address = "D:\\项目\\DP+\\实施部署API\\Pre环境\\profile_apiTest_45407_002.txt"
            jar_path = "D:\\downloads\\encrypt-1.0-SNAPSHOT(2).jar"
            with open(plain_address, 'w') as profile_file:
                profile_file.writelines(lines)

            data = run_java_jar(plain_address, public_key_address, Private_key_address,passwords,output_address, jar_path)
            print(data)
        '''
            data写入到D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx文件的B2单元格内，第二次循环写入到B3,第三次循环写入到B4，依次类推
            
        '''
        # 将数据写入 Excel 文件的不同行
        wb = load_workbook('D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx')
        ws = wb.active
        ws.cell(row=_ + 2, column=1, value=str(first_element))
        ws.cell(row=_ + 2, column=2, value=data)
        wb.save('D:\\项目\\DP+\\实施部署API\\Pre环境\\profile.xlsx')
        # 将修改后的内容写回文件
        with open(file_path, 'w') as file:
            file.writelines(lines)

file_path = 'D:\\项目\\DP+\\实施部署API\\Pre环境\\明文.txt'

num_iterations = int(input("请输入你需要多少组测试数据："))
print("你输入的是：", num_iterations)
input("按任意键继续...")

modify_file_content(file_path, num_iterations)



